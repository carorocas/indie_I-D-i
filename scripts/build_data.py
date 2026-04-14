"""Build static JSON files for the territorial I+D+i index viewer.

The frontend never reads the Excel workbook. Run this script after updating
CALCULO_resumenV2.xlsx to regenerate public/data/*.json.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
import urllib.request
from collections import defaultdict
from datetime import datetime, timezone
from difflib import get_close_matches
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "CALCULO_resumenV2.xlsx"
DATA_DIR = ROOT / "public" / "data"
GEOJSON_PATH = DATA_DIR / "colombia_departamentos.geojson"
GEOJSON_SOURCE_URL = (
    "https://raw.githubusercontent.com/caticoa3/colombia_mapa/master/"
    "co_2018_MGN_DPTO_POLITICO.geojson"
)

NOMBRE_INDICE = "Indice de Capacidades de I+D+i para el control integral del cancer"
ANIO = 2026


DEPARTAMENTO_ALIASES = {
    "bogota dc": "Bogotá D.C.",
    "bogota d c": "Bogotá D.C.",
    "bogota d.c": "Bogotá D.C.",
    "bogota d c.": "Bogotá D.C.",
    "bogota distrito capital": "Bogotá D.C.",
    "archipielago de san andres providencia y santa catalina": "San Andrés y Providencia",
    "san andres providencia y santa catalina": "San Andrés y Providencia",
    "san andres y providencia": "San Andrés y Providencia",
    "quindio": "Quindío",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def key(value: Any) -> str:
    text = clean_text(value).lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def numeric(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def canonical_department(value: Any, valid_departments: dict[str, str] | None = None) -> str:
    label = clean_text(value)
    label_key = key(label)
    if valid_departments and label_key in valid_departments:
        return valid_departments[label_key]
    if label_key in DEPARTAMENTO_ALIASES:
        return DEPARTAMENTO_ALIASES[label_key]
    return label


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def ranked(records: list[dict[str, Any]], value_name: str = "valor") -> list[dict[str, Any]]:
    ordered = sorted(
        records,
        key=lambda item: (item.get(value_name) is None, -(item.get(value_name) or 0), item["departamento"]),
    )
    for index, item in enumerate(ordered, start=1):
        item["ranking"] = index
        item["orden"] = index
    return ordered


def row_values(ws, min_row: int = 1):
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        yield [clean_text(cell) if isinstance(cell, str) else cell for cell in row]


def departments_from_header(header: list[Any], start_index: int) -> list[str]:
    return [canonical_department(cell) for cell in header[start_index:] if clean_text(cell)]


def canonical_subpilar(
    pilar: str,
    subpilar: str,
    known_by_pilar: dict[str, list[str]],
) -> str:
    candidates = known_by_pilar.get(pilar, [])
    if subpilar in candidates:
        return subpilar

    candidate_keys = {key(candidate): candidate for candidate in candidates}
    sub_key = key(subpilar)
    if sub_key in candidate_keys:
        return candidate_keys[sub_key]

    matches = get_close_matches(sub_key, candidate_keys.keys(), n=1, cutoff=0.82)
    if matches:
        return candidate_keys[matches[0]]
    return subpilar


def build_geojson(departments: list[str]) -> dict[str, Any] | None:
    valid = {key(name): name for name in departments}
    if not GEOJSON_PATH.exists():
        req = urllib.request.Request(GEOJSON_SOURCE_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=60) as response:
            raw = json.load(response)
    else:
        raw = json.loads(GEOJSON_PATH.read_text(encoding="utf-8"))

    features = []
    for feature in raw.get("features", []):
        props = feature.get("properties", {})
        raw_name = props.get("DPTO_CNMBR") or props.get("departamento") or props.get("name")
        department = canonical_department(raw_name, valid)
        if key(department) not in valid:
            continue
        features.append(
            {
                "type": "Feature",
                "properties": {
                    "codigo": clean_text(props.get("DPTO_CCDGO") or props.get("codigo")),
                    "departamento": department,
                    "fuente_nombre": clean_text(raw_name),
                },
                "geometry": feature.get("geometry"),
            }
        )

    return {
        "type": "FeatureCollection",
        "name": "colombia_departamentos",
        "source": "DANE 2018, via caticoa3/colombia_mapa",
        "features": sorted(features, key=lambda item: item["properties"]["departamento"]),
    }


def build(excel_path: Path) -> None:
    if not excel_path.exists():
        raise FileNotFoundError(f"No existe el Excel fuente: {excel_path}")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(excel_path, data_only=True, read_only=True)

    required = {"indicadores", "subpilares", "pilares", "total"}
    missing = required.difference(wb.sheetnames)
    if missing:
        raise ValueError(f"Faltan hojas requeridas: {', '.join(sorted(missing))}")

    # indicadores
    ws_ind = wb["indicadores"]
    ind_rows = list(row_values(ws_ind))
    ind_header = ind_rows[0]
    departments = departments_from_header(ind_header, 3)

    indicadores_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    indicators_by_subpilar: dict[tuple[str, str], list[str]] = defaultdict(list)
    subpilares_by_pilar: dict[str, list[str]] = defaultdict(list)

    for row in ind_rows[1:]:
        pilar = clean_text(row[0])
        subpilar = clean_text(row[1])
        indicador = clean_text(row[2])
        if not (pilar and subpilar and indicador):
            continue
        if subpilar not in subpilares_by_pilar[pilar]:
            subpilares_by_pilar[pilar].append(subpilar)
        if indicador not in indicators_by_subpilar[(pilar, subpilar)]:
            indicators_by_subpilar[(pilar, subpilar)].append(indicador)

        base_records = []
        for department, value in zip(departments, row[3:]):
            base_records.append(
                {
                    "pilar": pilar,
                    "subpilar": subpilar,
                    "indicador": indicador,
                    "departamento": department,
                    "valor": numeric(value),
                }
            )
        indicadores_by_name[indicador].extend(ranked(base_records))

    indicadores = []
    for indicador in sorted(indicadores_by_name):
        indicadores.extend(indicadores_by_name[indicador])

    known_subpilares = {pilar: list(items) for pilar, items in subpilares_by_pilar.items()}

    # subpilares
    ws_sub = wb["subpilares"]
    sub_rows = list(row_values(ws_sub))
    sub_header = sub_rows[0]
    sub_departments = departments_from_header(sub_header, 2)
    subpilares_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in sub_rows[1:]:
        pilar = clean_text(row[0])
        subpilar = canonical_subpilar(pilar, clean_text(row[1]), known_subpilares)
        if not (pilar and subpilar):
            continue
        indicadores_componentes = indicators_by_subpilar.get((pilar, subpilar), [])
        base_records = []
        for department, value in zip(sub_departments, row[2:]):
            base_records.append(
                {
                    "pilar": pilar,
                    "subpilar": subpilar,
                    "departamento": department,
                    "valor": numeric(value),
                    "indicadores_componentes": indicadores_componentes,
                }
            )
        subpilares_by_name[subpilar].extend(ranked(base_records))

    subpilares = []
    for subpilar in sorted(subpilares_by_name):
        subpilares.extend(subpilares_by_name[subpilar])

    # pilares
    ws_pil = wb["pilares"]
    pil_rows = list(row_values(ws_pil))
    pil_header = pil_rows[0]
    pil_departments = departments_from_header(pil_header, 1)
    pilares_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in pil_rows[1:]:
        pilar = clean_text(row[0])
        if not pilar:
            continue
        base_records = []
        for department, value in zip(pil_departments, row[1:]):
            base_records.append(
                {
                    "pilar": pilar,
                    "departamento": department,
                    "valor": numeric(value),
                    "subpilares_componentes": subpilares_by_pilar.get(pilar, []),
                }
            )
        pilares_by_name[pilar].extend(ranked(base_records))

    pilares = []
    for pilar in sorted(pilares_by_name):
        pilares.extend(pilares_by_name[pilar])

    pilar_lookup = {(item["departamento"], item["pilar"]): item["valor"] for item in pilares}
    subpilar_lookup = {
        (item["departamento"], item["pilar"], item["subpilar"]): item["valor"] for item in subpilares
    }

    # total
    ws_total = wb["total"]
    total_rows = list(row_values(ws_total))[1:]
    total_base = []
    for row in total_rows:
        department = canonical_department(row[0])
        if not department:
            continue
        total_base.append(
            {
                "departamento": department,
                "valor_total": numeric(row[1]),
                "pilares": [
                    {"nombre": pilar, "valor": pilar_lookup.get((department, pilar))}
                    for pilar in sorted(pilares_by_name)
                ],
                "subpilares": [
                    {
                        "nombre": subpilar,
                        "pilar": pilar,
                        "valor": subpilar_lookup.get((department, pilar, subpilar)),
                    }
                    for pilar in sorted(subpilares_by_pilar)
                    for subpilar in subpilares_by_pilar[pilar]
                ],
            }
        )
    total = ranked(total_base, "valor_total")

    metadata = {
        "nombre_indice": NOMBRE_INDICE,
        "anio": ANIO,
        "cobertura": "32 departamentos de Colombia y Bogotá D.C.",
        "departamentos": departments,
        "fuente_excel": excel_path.name,
        "generado_en": datetime.now(timezone.utc).isoformat(),
        "pilares": [
            {
                "nombre": pilar,
                "subpilares": [
                    {
                        "nombre": subpilar,
                        "indicadores": indicators_by_subpilar.get((pilar, subpilar), []),
                    }
                    for subpilar in subpilares_by_pilar[pilar]
                ],
            }
            for pilar in subpilares_by_pilar
        ],
        "subpilares": [
            {"nombre": subpilar, "pilar": pilar}
            for pilar, items in subpilares_by_pilar.items()
            for subpilar in items
        ],
        "indicadores": [
            {"nombre": indicador, "subpilar": subpilar, "pilar": pilar}
            for (pilar, subpilar), items in indicators_by_subpilar.items()
            for indicador in items
        ],
        "relaciones": {
            "pilar_subpilares": subpilares_by_pilar,
            "subpilar_indicadores": {
                f"{pilar} > {subpilar}": items
                for (pilar, subpilar), items in indicators_by_subpilar.items()
            },
        },
        "decisiones_cartograficas": [
            "Se usa una capa departamental DANE 2018 en GeoJSON como archivo estático.",
            "Bogotá D.C. está representada como entidad separada en la capa fuente.",
            "La entidad cartográfica Archipiélago de San Andrés, Providencia y Santa Catalina se etiqueta como San Andrés y Providencia para coincidir con el Excel.",
        ],
        "decisiones_limpieza": [
            "Los encabezados y nombres se limpian eliminando espacios redundantes.",
            "Las variantes de Bogotá D.C. y San Andrés se normalizan para el frontend.",
            "La variante 'Productos de nuevo conocimeinto' se alinea con 'Productos de nuevo conocimiento' por coincidencia textual dentro del mismo pilar.",
        ],
    }

    geojson = build_geojson(departments)
    if geojson:
        write_json(GEOJSON_PATH, geojson)

    write_json(DATA_DIR / "indicadores.json", indicadores)
    write_json(DATA_DIR / "subpilares.json", subpilares)
    write_json(DATA_DIR / "pilares.json", pilares)
    write_json(DATA_DIR / "total.json", total)
    write_json(DATA_DIR / "metadata.json", metadata)

    print(f"Datos generados en {DATA_DIR}")
    print(f"Indicadores: {len(indicadores)} registros")
    print(f"Subpilares: {len(subpilares)} registros")
    print(f"Pilares: {len(pilares)} registros")
    print(f"Total: {len(total)} registros")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Ruta al archivo Excel fuente")
    args = parser.parse_args()
    build(args.excel.resolve())


if __name__ == "__main__":
    main()
